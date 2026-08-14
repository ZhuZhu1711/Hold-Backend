"""报表 xlsx 导出：无数据库的小函数测试。"""
from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from app.controllers.dispose_ctrl import circulation_export_row
from app.controllers.engineer_ctrl import engineer_holding_export_row
from app.controllers.hold_report_ctrl import (
    HOLDING_EXPORT_HEADERS,
    _parse_page,
    hold_history_table,
    holding_export_row,
)
from app.controllers.production_ctrl import production_holding_export_row
from app.controllers.quality_ctrl import quality_export_row
from app.utils.excel_export import (
    EXPORT_MAX_ROWS,
    build_xlsx,
    cell,
    from_page_payload,
)


class CellAndWorkbookTest(unittest.TestCase):
    def test_cell(self):
        self.assertEqual(cell(None), '')
        self.assertEqual(cell(True), '是')
        self.assertEqual(cell(False), '否')
        self.assertEqual(cell(12), 12)
        self.assertEqual(cell('abc'), 'abc')

    def test_build_xlsx(self):
        content = build_xlsx('测试表', ['A', 'B'], [[1, None], [True, 'x']])
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.title, '测试表')
        self.assertEqual([c.value for c in ws[1]], ['A', 'B'])
        self.assertEqual([c.value for c in ws[2]], [1, None])
        self.assertEqual([c.value for c in ws[3]], ['是', 'x'])

    def test_from_page_payload_truncate_note(self):
        ok, msg, content = from_page_payload(
            True,
            '获取成功',
            {'items': [{'ID': 1}], 'total': 3},
            ['ID'],
            lambda it: [it.get('ID')],
            'Sheet',
        )
        self.assertTrue(ok)
        self.assertIn('共 3 条', msg)
        self.assertIn('已导出前 1 条', msg)
        self.assertTrue(content)

    def test_from_page_payload_fail(self):
        ok, msg, content = from_page_payload(False, '无效参数', None, ['A'], lambda x: x, 'S')
        self.assertFalse(ok)
        self.assertEqual(msg, '无效参数')
        self.assertIsNone(content)


class RowMapperTest(unittest.TestCase):
    def test_holding_export_row(self):
        row = holding_export_row({
            'ID': 9,
            'RECORD_TYPE_NAME': 'FT异常反馈单',
            'PRODUCT_ID': 'P1',
            'STATION': 'S1',
            'EQUIP_ID': 'E1',
            'LOT_ID': 'L1',
            'WAFER_ID': 'W1',
            'HOLD_CODE': '023',
            'HOLD_REASON': 'yield',
            'GRADE_NUM_DISPLAY': 'F:1',
            'CURRENT_OWNER_NAME': 'eng',
            'HOLD_DTTM': '2026-08-01 00:00:00',
            'IS_CLOSED': False,
        })
        self.assertEqual(len(row), len(HOLDING_EXPORT_HEADERS))
        self.assertEqual(row[0], 9)
        self.assertEqual(row[-1], 'Holding')
        self.assertEqual(holding_export_row({'IS_CLOSED': True})[-1], '已关闭')

    def test_engineer_holding_export_row(self):
        row = engineer_holding_export_row({
            'ID': 1,
            'CAN_DISPOSE': True,
        })
        self.assertEqual(row[-1], '是')
        self.assertEqual(engineer_holding_export_row({'CAN_DISPOSE': False})[-1], '否')

    def test_circulation_export_row(self):
        row = circulation_export_row({
            'ID': 10,
            'HOLD_RECORD_ID': 20,
            'PRODUCT_ID': 'P1',
            'DISPOSE_LABEL': '放行',
            'DISPOSED_OWNER_NAME': 'eng',
            'NEXT_OWNER_ID': 181,
        })
        self.assertEqual(row[0], 10)
        self.assertEqual(row[1], 20)
        self.assertEqual(row[6], '放行')
        self.assertEqual(row[8], 181)

    def test_quality_export_row(self):
        row = quality_export_row({
            'DISPOSE_DTTM': '2026-08-01 12:00:00',
            'PRODUCT_ID': 'P1',
            'RECORD_TYPE_NAME': 'FVI异常反馈单',
            'DISPOSE_LABEL': '降级',
            'HOLD_RECORD_ID': 33,
        })
        self.assertEqual(row[0], '2026-08-01 12:00:00')
        self.assertEqual(row[4], '降级')
        self.assertEqual(row[-1], 33)

    def test_production_holding_export_row(self):
        row = production_holding_export_row({
            'ID': 5,
            'LAST_DISPOSE_LABEL': '放行',
            'GRADE_NUM_DISPLAY': 'F:2',
        })
        self.assertEqual(row[0], 5)
        self.assertEqual(row[9], '放行')
        self.assertEqual(row[-2], 'F:2')
        self.assertEqual(row[-1], '否')
        self.assertEqual(
            production_holding_export_row({'PENDING_SAMPLE_RETAIN': True})[-1],
            '是',
        )


class HistoryTableTest(unittest.TestCase):
    def test_hold_history_table(self):
        headers, rows = hold_history_table({
            'labels': ['2026-08-01', '2026-08-02'],
            'series': [
                {'name': 'FT异常反馈单', 'values': [1, 2], 'total': 3},
                {'name': 'FVI异常反馈单', 'values': [0, 1], 'total': 1},
                {'name': 'WLT异常反馈单', 'values': [4, 0], 'total': 4},
            ],
            'total': 8,
        })
        self.assertEqual(headers, ['日期', 'FT异常反馈单', 'FVI异常反馈单', 'WLT异常反馈单', '合计'])
        self.assertEqual(rows[0], ['2026-08-01', 1, 0, 4, 5])
        self.assertEqual(rows[1], ['2026-08-02', 2, 1, 0, 3])
        self.assertEqual(rows[2], ['合计', 3, 1, 4, 8])


class ParsePageExportTest(unittest.TestCase):
    def test_export_page_size_not_capped_at_200(self):
        page, page_size, offset = _parse_page(1, EXPORT_MAX_ROWS, max_page_size=EXPORT_MAX_ROWS)
        self.assertEqual(page, 1)
        self.assertEqual(page_size, EXPORT_MAX_ROWS)
        self.assertEqual(offset, 0)

    def test_default_still_capped(self):
        _, page_size, _ = _parse_page(1, 5000)
        self.assertEqual(page_size, 200)


if __name__ == '__main__':
    unittest.main()
