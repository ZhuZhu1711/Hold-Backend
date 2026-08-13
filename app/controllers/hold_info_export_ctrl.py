"""
按型号 + 时间从 FT_HOLD_RECORD 导出 FT ATE Hold Lot 表（原 HoldInfoExport 脚本）。

逻辑：
  1. 主数据取 FT_HOLD_RECORD（合批任务已合并，不再对 HOLD_INFO 做 1 分钟窗口）
  2. ROUTE_ID 含 MP → PRODUCTION，含 ENG → ENG
  3. 合批片（LOT/WAFER 后缀位数>2）查 SPLIT_MERGE_HISTORY 还原源 lot / wafer
  4. 补 TEST_WAFER.TEST_PROGRAM
"""
from datetime import datetime
from io import BytesIO
import os
import re

import oracledb
from flask import current_app
from openpyxl import Workbook, load_workbook

from app.config import Config
from app.utils.database_util import (
    USER, PWD, DSN, expand_display_wafer_ids, is_merged_wafer_id, logger,
)

_ALLOWED_HOLD_RECORD_TABLES = {'FT_HOLD_RECORD'}

EXPORT_MAX_ROWS = 5000
PREVIEW_MAX_ROWS = 100

DEFAULTS = {
    'sub_customer': 'GC',
    'package_type': 'CIS',
    'factory': 'GC',
    'area': 'FT',
    'stage': 'FA',
    'tester': 'ANY',
}

EXCEL_HEADERS = [
    'SubCustomer', 'Product', 'Device', 'PackageType', 'Factory', 'Area',
    'RouteType', 'LotId', 'CustLotId', 'WaferId', '', 'Stage',
    'HoldWaferList', 'WaferList', 'HoldReason', 'Tester', '', '',
    'HoldDateTime', '', 'TestProgram', '', '', '', '', 'HoldDateTime2',
]

_DTTM_RE = re.compile(
    r'^(\d{4}-\d{2}-\d{2})(?:[ T](\d{2}:\d{2})(?::(\d{2}))?)?$'
)


def _connect():
    return oracledb.connect(user=USER, password=PWD, dsn=DSN)


def _record_table():
    name = (getattr(Config, 'HOLD_RECORD_TABLE', None) or 'FT_HOLD_RECORD').upper()
    if name not in _ALLOWED_HOLD_RECORD_TABLES:
        raise ValueError(f'非法 HOLD_RECORD 表名: {name}')
    return name


def normalize_dttm(raw, is_end=False):
    """
    接受 YYYY-MM-DD / YYYY-MM-DD HH:MM / YYYY-MM-DDTHH:MM[:SS]。
    仅日期时：开始补 00:00:00，结束补 23:59:59。
    成功返回 'YYYY-MM-DD HH:MM:SS'，失败返回 None。
    """
    text = (raw or '').strip().replace('T', ' ')
    if not text:
        return None
    m = _DTTM_RE.match(text)
    if not m:
        return None
    day, hm, sec = m.group(1), m.group(2), m.group(3)
    if not hm:
        hm = '23:59' if is_end else '00:00'
        sec = '59' if is_end else '00'
    elif not sec:
        sec = '59' if is_end else '00'
    return f'{day} {hm}:{sec}'


def format_wafer_list(wafer_list):
    """LOT:#01#02/LOT2:#03"""
    groups = []
    current_lot = None
    current_nos = []
    for wafer in sorted(str(w).strip() for w in wafer_list if w):
        if '-' in wafer:
            lot = wafer.split('-')[0]
            wafer_no = wafer.split('-')[-1]
        else:
            lot, wafer_no = wafer, ''
        if lot != current_lot:
            if current_lot is not None:
                groups.append(current_lot + ':' + ''.join('#' + n for n in current_nos))
            current_lot = lot
            current_nos = [wafer_no] if wafer_no else []
        elif wafer_no:
            current_nos.append(wafer_no)
    if current_lot is not None:
        groups.append(current_lot + ':' + ''.join('#' + n for n in current_nos))
    return '/'.join(groups)


def _query_hold_records(cursor, product_id, start_dttm, end_dttm, lot_id=''):
    record_table = _record_table()
    sql = f"""
        SELECT
            r.WAFER_ID,
            r.PRODUCT_ID,
            r.EQUIP_ID,
            TO_CHAR(r.HOLD_DTTM, 'YYYY-MM-DD HH24:MI:SS') AS HOLD_DATETIME,
            r.HOLD_REASON AS REASON,
            CASE
                WHEN r.ROUTE_ID LIKE '%MP%' THEN 'PRODUCTION'
                WHEN r.ROUTE_ID LIKE '%ENG%' THEN 'ENG'
            END AS ROUTE_TYPE,
            r.LOT_ID
        FROM {record_table} r
        WHERE r.PRODUCT_ID = :device
          AND r.HOLD_DTTM >= TO_DATE(:start_dttm, 'YYYY-MM-DD HH24:MI:SS')
          AND r.HOLD_DTTM <= TO_DATE(:end_dttm, 'YYYY-MM-DD HH24:MI:SS')
    """
    params = {
        'start_dttm': start_dttm,
        'end_dttm': end_dttm,
        'device': product_id,
    }
    if lot_id:
        sql += " AND UPPER(r.LOT_ID) LIKE UPPER(:lot_id)"
        params['lot_id'] = f'%{lot_id}%'
    sql += " ORDER BY r.WAFER_ID, r.HOLD_DTTM, r.ID"
    cursor.execute(sql, params)
    return cursor.fetchall()


def _query_merge_sources(cursor, wafer_id, cache):
    if wafer_id in cache:
        return cache[wafer_id]
    cursor.execute(
        """
        SELECT source_lot_id
        FROM mesprod.SPLIT_MERGE_HISTORY@MES16019 s
        WHERE s.TARGET_LOT_ID = :wafer_id
        ORDER BY source_lot_id ASC
        """,
        {'wafer_id': wafer_id},
    )
    sources = [str(r[0]).strip() for r in cursor.fetchall() if r and r[0]]
    cache[wafer_id] = sources
    return sources


def _query_hold_wafers_by_lot(cursor, lot, cache):
    if lot in cache:
        return cache[lot]
    record_table = _record_table()
    cursor.execute(
        f"""
        SELECT LOT_ID, WAFER_ID
        FROM {record_table}
        WHERE LOT_ID LIKE :pat OR WAFER_ID LIKE :pat
        """,
        {'pat': f'{lot}%'},
    )
    wafers = set()
    for lot_id, wafer_id in cursor.fetchall():
        for w in expand_display_wafer_ids(wafer_id, lot_id):
            if w:
                wafers.add(w)
    cache[lot] = wafers
    return wafers


def _query_test_program(cursor, wafer_id, route, cache):
    key = (wafer_id, route or '')
    if key in cache:
        return cache[key]
    if not wafer_id or not route:
        cache[key] = ''
        return ''
    cursor.execute(
        """
        SELECT TEST_PROGRAM
        FROM TEST_WAFER
        WHERE WAFER_ID = :wafer_id AND ROUTE LIKE :route_pat
        ORDER BY ID DESC
        """,
        {'wafer_id': wafer_id, 'route_pat': f'%{route}%'},
    )
    row = cursor.fetchone()
    value = str(row[0]).strip() if row and row[0] else ''
    cache[key] = value
    return value


def _merge_target(lot_id, wafer_id, expanded):
    if is_merged_wafer_id(lot_id):
        return lot_id
    if is_merged_wafer_id(wafer_id):
        return wafer_id
    for w in expanded:
        if is_merged_wafer_id(w):
            return w
    return None


def _process_hold(cursor, hold, caches, extras):
    wafer_id = str(hold[0] or '').strip()
    product_id = str(hold[1] or '').strip()
    hold_date = hold[3]
    reason = hold[4] or ''
    route = hold[5]
    lot_id = str(hold[6] or '').strip()

    product = product_id.split('-')[0] if product_id else ''
    expanded = expand_display_wafer_ids(wafer_id, lot_id)
    lookup_wafer = expanded[0] if expanded else wafer_id
    test_pro = _query_test_program(cursor, lookup_wafer, route, caches['test_program'])

    source_lots_set = set()
    source_wafers_set = set()
    merge_key = _merge_target(lot_id, wafer_id, expanded)
    if merge_key:
        for src in _query_merge_sources(cursor, merge_key, caches['merge']):
            source_wafers_set.add(src)
            source_lots_set.add(src.split('-')[0] if '-' in src else src)
    if not source_wafers_set:
        source_wafers_set.update(expanded or ([wafer_id] if wafer_id else []))
        if lot_id:
            source_lots_set.add(lot_id.split('-')[0] if '-' in lot_id else lot_id)
        elif wafer_id and '-' in wafer_id and not wafer_id.startswith('#'):
            source_lots_set.add(wafer_id.split('-')[0])

    hold_wafers_set = set(expanded)
    for lot in list(source_lots_set):
        for wafer in _query_hold_wafers_by_lot(cursor, lot, caches['hold_wafers']):
            if wafer in source_wafers_set:
                hold_wafers_set.add(wafer)
    if not hold_wafers_set:
        hold_wafers_set.update(source_wafers_set)

    source_lots = sorted(source_lots_set)
    source_wafers = sorted(source_wafers_set)
    hold_wafers = sorted(hold_wafers_set)
    lot_str = '/'.join(source_lots)
    hold_wafer_list_str = format_wafer_list(hold_wafers)
    wafer_list_str = format_wafer_list(source_wafers)
    if not hold_wafer_list_str:
        hold_wafer_list_str = wafer_list_str

    row = [
        extras.get('sub_customer') or DEFAULTS['sub_customer'],
        product,
        product_id,
        extras.get('package_type') or DEFAULTS['package_type'],
        extras.get('factory') or DEFAULTS['factory'],
        extras.get('area') or DEFAULTS['area'],
        route or '',
        lot_str,
        lot_str,
        wafer_id,
        None,
        extras.get('stage') or DEFAULTS['stage'],
        hold_wafer_list_str,
        wafer_list_str,
        reason,
        extras.get('tester') or DEFAULTS['tester'],
        None,
        None,
        hold_date,
        None,
        test_pro,
        None,
        None,
        None,
        None,
        hold_date,
    ]
    preview = {
        'WAFER_ID': wafer_id,
        'PRODUCT_ID': product_id,
        'EQUIP_ID': str(hold[2] or '').strip(),
        'HOLD_DATETIME': hold_date,
        'REASON': reason,
        'ROUTE_TYPE': route or '',
        'LOT_STR': lot_str,
        'HOLD_WAFER_LIST': hold_wafer_list_str,
        'WAFER_LIST': wafer_list_str,
        'TEST_PROGRAM': test_pro,
    }
    return row, preview


def _parse_filters(product_id, start_dttm, end_dttm, lot_id=''):
    product_id = (product_id or '').strip()
    lot_id = (lot_id or '').strip()
    start = normalize_dttm(start_dttm, is_end=False)
    end = normalize_dttm(end_dttm, is_end=True)
    if not product_id:
        return False, '请指定型号 PRODUCT_ID', None
    if not start or not end:
        return False, '请指定有效的开始/结束时间', None
    try:
        start_dt = datetime.strptime(start, '%Y-%m-%d %H:%M:%S')
        end_dt = datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        return False, '时间格式无效', None
    if end_dt < start_dt:
        return False, '结束时间不能早于开始时间', None
    return True, 'ok', (product_id, start, end, lot_id)


def _collect_rows(product_id, start_dttm, end_dttm, extras, limit, lot_id=''):
    conn = _connect()
    try:
        cursor = conn.cursor()
        holds = _query_hold_records(
            cursor, product_id, start_dttm, end_dttm, lot_id=lot_id,
        )
        total = len(holds)
        truncated = total > limit
        holds = holds[:limit]
        caches = {'merge': {}, 'hold_wafers': {}, 'test_program': {}}
        excel_rows = []
        preview_rows = []
        for hold in holds:
            excel_row, preview = _process_hold(cursor, hold, caches, extras)
            excel_rows.append(excel_row)
            preview_rows.append(preview)
        return True, 'ok', {
            'excel_rows': excel_rows,
            'preview_rows': preview_rows,
            'total': total,
            'truncated': truncated,
            'product_id': product_id,
            'lot_id': lot_id,
            'start_dttm': start_dttm,
            'end_dttm': end_dttm,
        }
    except Exception as e:
        logger.error(f'Hold Info 导出失败: {e}', exc_info=True)
        return False, f'导出失败: {e}', None
    finally:
        conn.close()


def preview_hold_info_export(product_id, start_dttm, end_dttm, lot_id='', **extras):
    ok, msg, parsed = _parse_filters(product_id, start_dttm, end_dttm, lot_id=lot_id)
    if not ok:
        return False, msg, None
    product_id, start, end, lot_id = parsed
    ok, msg, payload = _collect_rows(
        product_id, start, end, extras, PREVIEW_MAX_ROWS, lot_id=lot_id,
    )
    if not ok:
        return False, msg, None
    return True, '查询成功', {
        'items': payload['preview_rows'],
        'total': payload['total'],
        'shown': len(payload['preview_rows']),
        'truncated': payload['truncated'],
        'product_id': product_id,
        'lot_id': lot_id,
        'start_dttm': start,
        'end_dttm': end,
    }


def _find_template():
    name = 'FT_ATE_HOLD_LOT.xlsx'
    root = current_app.root_path
    backend_root = os.path.dirname(root)
    workspace = os.path.dirname(backend_root)
    candidates = [
        os.path.join(root, 'static', 'templates', name),
        os.path.join(workspace, 'HoldInfoExport', 'template', name),
    ]
    for path in candidates:
        if os.path.isfile(path):
            return path
    return None


def export_hold_info_xlsx(product_id, start_dttm, end_dttm, lot_id='', **extras):
    ok, msg, parsed = _parse_filters(product_id, start_dttm, end_dttm, lot_id=lot_id)
    if not ok:
        return False, msg, None
    product_id, start, end, lot_id = parsed
    ok, msg, payload = _collect_rows(
        product_id, start, end, extras, EXPORT_MAX_ROWS, lot_id=lot_id,
    )
    if not ok:
        return False, msg, None

    template = _find_template()
    if template:
        wb = load_workbook(template)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'HoldInfo'
        ws.append(EXCEL_HEADERS)

    for row in payload['excel_rows']:
        ws.append(row)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    note = f"导出 {len(payload['excel_rows'])} 条"
    if payload['truncated']:
        note = (
            f"共 {payload['total']} 条，已导出前 {len(payload['excel_rows'])} 条"
        )
    return True, note, bio.getvalue()
